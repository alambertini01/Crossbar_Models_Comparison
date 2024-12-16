import random
import datetime
import pytz
import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.colors import to_hex, to_rgb, LinearSegmentedColormap
from matplotlib.cm import ScalarMappable
from matplotlib.patches import Patch
from mpl_toolkits.axes_grid1.inset_locator import inset_axes
from math import pi
import time
import tkinter as tk
from tkinter import ttk
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from sklearn.linear_model import LinearRegression
#Import NonLinear functions
from CrossbarModels.Functions.NonLinear import resistance_array_to_x
from CrossbarModels.Functions.NonLinear import calculate_resistance
#Import the Crossbar Models
from CrossbarModels.Crossbar_Models import *
from CrossbarModels.Crossbar_Models_pytorch import gamma_model, dmr_model, jeong_model, crosssim_model



# Use this script to compare the performance (time and accuracy) of different parasitc resistance corsbar models


############################ PARAMETERS ##############################

# Dimensions of the crossbar
input,output = (128,128)

# Initialize each model instance
Models = [
    JeongModel("Jeong"),
    JeongModel_avg("Jeong_avg"),
    JeongModel_avgv2("Jeong_torch"),
    IdealModel("Ideal"),
    DMRModel("DMR_old"),
    DMRModel_acc("DMR_acc"),
    DMRModel_new("DMR"),
    DMRModel_new("DMR_torch"),
    GammaModel("Gamma_torch"),
    GammaModel_acc("Gamma_acc_v1"),
    GammaModel_acc_v2("Gamma"),
    CrossSimModel("CrossSim_ref"),
    CrossSimModel("CrossSim1",Verr_th=0.5),
    CrossSimModel("CrossSim2",Verr_th=1e-1),
    CrossSimModel("CrossSim3",Verr_th=1e-2),
    CrossSimModel("CrossSim4",Verr_th=1e-3),
    CrossSimModel("CrossSim5",Verr_th=1e-4),
    CrossSimModel("CrossSim6",Verr_th=1e-5),
    CrossSimModel("CrossSim7",Verr_th=1e-6),
    CrossSimModel("CrossSim8",Verr_th=1e-7),
    CrossSimModel("CrossSim9",Verr_th=1e-8),
    CrossSimModel("CrossSim10",Verr_th=1e-9),
    CrossSimModel("CrossSim_torch",Verr_th=1e-10),
    LTSpiceModel("LTSpice"),
    NgSpiceModel("NgSpice"),
    NgSpiceNonLinearModel("NgSpiceNonLinear"),
    MemtorchModelCpp("Memtorch"),
    MemtorchModelPython("Memtorch_python")
]

new_model_functions = {
    "Gamma_torch": gamma_model,
    "DMR_torch": dmr_model,
    "Jeong_torch": jeong_model,
    "CrossSim_torch" : crosssim_model
}
enabled_models = [ "Ideal","Jeong","DMR","Gamma"]
enabled_models = [ "Ideal","Jeong","DMR","Gamma","CrossSim2", "CrossSim3", "CrossSim4", "CrossSim5", "CrossSim6", "CrossSim7", "CrossSim8", "CrossSim9", "Memtorch", "NgSpice"]
# enabled_models = [model.name for model in Models]
# enabled_models = [ "Ideal","Jeong_avgv2","DMR","Gamma","CrossSim1","CrossSim2","CrossSim3","CrossSim4","CrossSim5","CrossSim6","CrossSim7","CrossSim8","CrossSim9","CrossSim10","Memtorch","NgSpice"]

reference_model =  "CrossSim10"

# Low resistance proggramming value
R_lrs = 1000
Rhrs_percentage=50
# parasitic resistance value
parasiticResistance = np.arange(0.2, 5, 0.2)
parasiticResistance = np.array([2])

# Memory window (ratio between Hrs and Lrs)
memoryWindow = np.arange(5, 101, 5)
memoryWindow = np.array([20])

# Input voltages parameters
v_On_percentage = 100
population = [0.5, 0.0]

# Metric type (2=Current*Times, 1=Current, 0=Voltage)
Metric_type = 1

# Variability parameters
v_flag = 1
v_size = 2





############################ INITIALIZATIONS ############################

enabled_models.append(reference_model)

memorySize = np.size(memoryWindow)
parasiticSize = np.size(parasiticResistance)
modelSize = len(enabled_models)
totalIterations = memorySize*parasiticSize*modelSize

# Tech 4 parameters
rho=3000
tox=5
s0=10.62
s=rho*tox/R_lrs
v_ds = 0.35
v_dx = 0.7

# Building the Resistance matrices
R = np.zeros((np.size(parasiticResistance),memorySize, v_size, input, output))
X = np.zeros((np.size(parasiticResistance),memorySize, v_size, input, output))
S = np.ones((np.size(parasiticResistance),memorySize, v_size, input, output))


# Generate Potential vector
Potential = np.random.choice(population, size=input, p=[v_On_percentage / 100, 1 - v_On_percentage / 100])

reference_index = enabled_models.index(reference_model)

# Output Currents (for each Parasitic Model)
output_currents = np.zeros((output, parasiticSize ,memorySize, modelSize))

# Voltage Drops (for each Parasitic Model)
V_a_matrix = np.tile(Potential, output)
voltage_drops = np.zeros((input,output, parasiticSize ,memorySize, modelSize))

# Initialize Metric arrays
Current_error = np.zeros((parasiticSize ,memorySize, modelSize-1))
Voltage_error = np.zeros((parasiticSize ,memorySize, modelSize-1))

# Initialize time measurements
simulation_times = np.zeros(modelSize)


# ########## Progress Bar ###########
def update_progress_bars(memory_progress, resistance_progress, variability_progress, m, z, v, memorySize, parasiticResistance, v_size):
    memory_progress['value'] = (m + 1) / memorySize * 100
    resistance_progress['value'] = (z + 1) / np.size(parasiticResistance) * 100
    variability_progress['value'] = (v + 1) / v_size * 100
    root.update_idletasks()
    root.update()            # Update the Tkinter event loop
# Create a tkinter window with 3 progress bars
root = tk.Tk()
root.title("Simulation Progress")
memory_progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
memory_progress.pack(pady=10)
memory_progress_label = tk.Label(root, text="Memory Window")
memory_progress_label.pack()
resistance_progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
resistance_progress.pack(pady=10)
resistance_progress_label = tk.Label(root, text="Parasitic Resistance")
resistance_progress_label.pack()
variability_progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
variability_progress.pack(pady=10)
variability_progress_label = tk.Label(root, text="Variability")
variability_progress_label.pack()
root.update_idletasks()




############################ MAIN LOOP ############################

for m in range(memorySize):
    update_progress_bars(memory_progress, resistance_progress, variability_progress, m, 0, 0, memorySize, parasiticResistance, v_size)
    x_thickness = resistance_array_to_x(np.array([R_lrs*memoryWindow[m]]))

    for z in range(np.size(parasiticResistance)):
        update_progress_bars(memory_progress, resistance_progress, variability_progress, m, z, 0, memorySize, parasiticResistance, v_size)

        for v in range(v_size):
            update_progress_bars(memory_progress, resistance_progress, variability_progress, m, z, v, memorySize, parasiticResistance, v_size)

            # Generate Resistance Matrix based on the parameters
            Rstate = np.random.choice([1, 0], size=(input, output), p=[Rhrs_percentage/100, 1 - Rhrs_percentage/100])
            # Rstate[5,10]=1
            X[z,m,v] = Rstate*x_thickness+10e-13
            S[z,m,v] = (1-Rstate)*s + Rstate*s0
            if v_flag:
                # barrier thickness variability for non linear case
                X[z,m,v] += (np.abs(np.random.randn(input,output)*v_dx/3))*Rstate
                S[z,m,v] += np.random.randn(input,output)*v_ds/3
                # Calculate the resistance with the variability R[z, m, v]
                R[z, m, v] = calculate_resistance(X[z,m,v], S[z,m,v])
            else:
                R[z,m,v] = Rstate*R_lrs*memoryWindow[m]+(1-Rstate)*R_lrs

            # Simulate the crossbar with each model
            for index, model in enumerate(Models):
                if model.name in enabled_models:
                    model_idx = enabled_models.index(model.name)
                    
                    # Prepare inputs for both old and new style models
                    # Old models use: model.calculate(R, parasiticResistance, Potential, ...)
                    # New models use: function(weight, x, parasiticResistance)
                    
                    # Convert R to weight = 1/R for new models
                    # R[z, m, v] is (input, output)
                    # For batch_size=1: weight should be (1, input, output)
                    weight_np = 1.0 / R[z, m, v]
                    weight_tensor = torch.tensor(weight_np, dtype=torch.float32) # shape (1, input, output)
                    
                    # Potential is of shape (input,). For batch_size=1, x should be (1, input)
                    x_np = Potential[np.newaxis, :]  # shape (1, input)
                    x_tensor = torch.tensor(x_np, dtype=torch.float32)
                    
                    # Check if model name corresponds to a new-style function
                    if model.name in new_model_functions:
                        # Call the new PyTorch function
                        start_time = time.perf_counter()
                        with torch.no_grad():
                            out_curr_torch = new_model_functions[model.name](weight_tensor, x_tensor, parasiticResistance[z])
                        end_time = time.perf_counter()
                        
                        # v_drops_torch: (1, input, output)
                        # out_curr_torch: (1, output)
                        # Convert back to numpy and remove batch dimension

                        out_curr_np = out_curr_torch.squeeze(0).numpy() # (output,)

                        output_currents[:, z, m, model_idx] = out_curr_np
                        
                        simulation_times[model_idx] += (end_time - start_time)/totalIterations
                    
                    else:
                        # Old model class with calculate function
                        # NonLinear_params needed only for NgSpiceNonLinear as shown in original code
                        NonLinear_params = {'X': X[z,m,v], 'S': S[z,m,v]} if model.name == 'NgSpiceNonLinear' else {'R_lrs': R_lrs, 'MW':memoryWindow[m]}
                        
                        start_time = time.perf_counter()
                        v_drop, out_curr = model.calculate(R[z, m, v], parasiticResistance[z], Potential, **NonLinear_params)
                        end_time = time.perf_counter()
                        
                        voltage_drops[:, :, z, m, model_idx] = v_drop
                        output_currents[:, z, m, model_idx] = out_curr
                        
                        simulation_times[model_idx] += (end_time - start_time)/totalIterations
            
            # Calculate the error metrics
            for index, model in enumerate(enabled_models[:-1]):
                # Compute Output Current Metric
                Current_error[z, m, index] += np.mean(np.abs(output_currents[:, z, m, reference_index] - output_currents[:, z, m, index] ) / output_currents[:, z, m, reference_index])*100/v_size
                # Compute Voltage Metric
                Voltage_error[z, m, index] += np.mean(np.abs((voltage_drops[:,:,z,m,reference_index] - voltage_drops[:,:,z,m,index]))/voltage_drops[:,:,z,m,reference_index])*100/v_size

root.destroy()  # Close the window when simulation is complete

if Metric_type==2:
    Metric = Current_error * simulation_times[np.newaxis, np.newaxis, :]
elif Metric_type==1:
    Metric = Current_error
else:
    Metric = Voltage_error

if "Ideal" in enabled_models:
    index = enabled_models.index("Ideal")
    # Divide simulation_times by the element at the found index
    normalized_simulation_times = simulation_times / simulation_times[index]
    if Metric_type==2:
        Metric = np.reciprocal(Current_error[:,:,:-1] * normalized_simulation_times[np.newaxis, np.newaxis, :-1])











###################### Plotting portion ##########################################################################



# Known color mapping
color_mapping = {
    "Jeong": "c",
    "DMR": "g",
    "Gamma": "r",
    "Ng": "pink",
    "CrossSim": "b",
    "Ideal": "black",
    "Memtorch": "orange",
    "LTSpice": "m"
}

# Generate the colors list
colors = [
    color_mapping[next((key for key in color_mapping if model.startswith(key)), None)] 
    if any(model.startswith(key) for key in color_mapping) 
    else "#{:06x}".format(random.randint(0, 0xFFFFFF))  # Generate a random color
    for model in enabled_models
]

markers = ['o', 's', 'D', '^', 'v', 'p']

plt.rcParams.update({
    'font.size': 14,
    'font.family': 'Arial',
    'axes.labelsize': 16,
    'axes.titlesize': 18,
    'legend.fontsize': 14,
    'xtick.labelsize': 14,
    'ytick.labelsize': 14
})

# Figures Selection
Simulation_times_plot = 1

Absolute_current_plots = 1
Relative_error_plots = 1

Voltage_drops_plot = 1
Voltage_drops_error_plot = 0

Metric_plot = 1
Mean_Metric_plot = 1
Metric_vs_Rpar = 1
Metric_vs_MW = 1
Winning_models_map = 1
print_table = 1
scatter_plot = 1
spider_plot = 1

Resistance_heatmap =0
open_folder = 1


x = np.arange(0, output)
# Results Folder
end = datetime.datetime.now(pytz.timezone('Europe/Rome'))
folder = 'Results/'+str(end.year)+ str(end.month)+  str(end.day) + '_'+ str(end.hour) +'_'+ str(end.minute)+"_"+str(input)+'x'+str(output)
if not (os.path.exists(folder)):
        os.makedirs(folder)
script_dir = os.path.dirname(os.path.abspath(__file__))
folder_path = os.path.join(script_dir, folder)

# Check there are not too many instances of plots
if memorySize>3 or parasiticSize>3:
    Absolute_current_plots = 0
    Relative_error_plots = 0
    Voltage_drops_plot = 0
    Voltage_drops_error_plot = 0
    Resistance_heatmap =0
else:
    Metric_plot=0
    Metric_vs_Rpar=0
    Metric_vs_MW=0
    print_table=0
    Winning_models_map = 0

# Different labels based on the used metric
if Metric_type:
    error_label = "Normalized Output Current Error (%)"
    if Metric_type==2:
        error_label = "Precision over Execution Time [1/s]"
else:
    error_label = "Normalized Voltage Drops Error (%)"
array_size_string = "Array Size: "+ str(input) + 'x' + str(output)




if Simulation_times_plot:
    plt.figure()
    if "Ideal" in enabled_models:
        # Create new lists/arrays without the "Ideal" element
        new_enabled_models = [model for i, model in enumerate(enabled_models) if i != index]
        new_simulation_times = np.delete(normalized_simulation_times, index)
        plt.title('Normalized Processing Time relative to the Ideal Model'+ '\n'+array_size_string)
        plt.ylabel('Normalized Time (log scale)')
        bars = plt.bar(new_enabled_models, new_simulation_times, color=[colors[(i+1) % len(colors)] for i in range(len(enabled_models))])
    else:
        plt.title('Processing Time for Each Model (Log Scale)'+ '\n'+array_size_string)
        plt.ylabel('Time (seconds, log scale)')
        bars = plt.bar(enabled_models, simulation_times, color=[colors[i % len(colors)] for i in range(len(enabled_models))])

    plt.xlabel('Model')
    plt.yscale('log')
    # Optionally rotate x-labels if they overlap
    plt.xticks(rotation=45, ha='right')
    # Adjust layout to prevent label cutoff
    plt.tight_layout()
    plt.savefig(folder + '/Figure_Simulation_times.png')
    plt.show()
    print("Simulation Times:",simulation_times)


# Output Currents plot
for z in range(np.size(parasiticResistance)):
    if Relative_error_plots:
        relative_Current = np.zeros(modelSize)
        # First figure: Relative error plots
        figure1, axis1 = plt.subplots(1, memorySize, figsize=(10, 5))  # Create a single row of subplots
        if memorySize == 1:
            axis1 = np.array([axis1])  # Ensure axis1[0] structure for single column case
        for m in range(memorySize):
            for index, model in enumerate(enabled_models):
                if model != reference_model:
                    # Calculate relative currents
                    relative_Current = np.abs(output_currents[:, z, m, reference_index] - output_currents[:, z, m, index] ) / output_currents[:, z, m, reference_index]
                    # Plot relative errors
                    axis1[m].plot(x, relative_Current * 100, color=colors[index % len(colors)], marker='.', label=(model))
                    axis1[m].set_title(f"Relative Error  (MW = {memoryWindow[m]}, Rpar = {parasiticResistance[z]} Ohm)"+ '\n'+array_size_string)
                    axis1[m].set(xlabel='jth bit line', ylabel='Normalized Output Current Error (%)')
                    axis1[m].legend()
        figure1.tight_layout()
        figure1.savefig(folder + f'/Figure_Relative_Current_Error_Rpar={z}.png')
        figure1.show()

    if Absolute_current_plots:
        # Second figure: Absolute current plots 
        figure2, axis2 = plt.subplots(1, memorySize, figsize=(10, 5))  # Create a single row of subplots
        if memorySize == 1:
            axis2 = np.array([axis2])  # Ensure axis2[0] structure for single column case
        for m in range(memorySize):
            for index, model in enumerate(enabled_models):
                axis2[m].plot(x, output_currents[:, z, m, index], color=colors[index % len(colors)], label=(model))
                axis2[m].set_title(f"Output currents  (MW = {memoryWindow[m]}, Rpar = {parasiticResistance[z]} Ohm)"+ '\n'+array_size_string)
                axis2[m].set(xlabel='jth bit line', ylabel='Current [A]')
                axis2[m].legend()
        figure2.tight_layout()
        figure2.savefig(folder + f'/Figure2_Absolute_Current{z}.png')
        figure2.show()


if Voltage_drops_plot:
    for m in range(memorySize):
        for z in range(np.size(parasiticResistance)):
            # Voltage drops comparison (Real vs Jeong vs DMR vs Gamma)
            Vmin = np.min(voltage_drops)
            # Create a figure with subplots side by side
            voltage_fig = plt.figure(figsize=(15, 10))
            for index, model in enumerate(enabled_models):
                # Plot heatmap for each model
                plt.subplot(2, round((modelSize+1)/2), index+1)  # 1 row, 3 columns, 1st subplot
                plt.imshow(voltage_drops[:,:,z,m,index], cmap='hot', interpolation='nearest',vmin=Vmin, vmax=1)
                plt.colorbar(label='Voltage Drop (V)')
                plt.title('Voltage Drop Heatmap ('+ model +')')
                plt.xlabel('Column Index (j)')
                plt.ylabel('Row Index (m)')
            plt.tight_layout()
            plt.savefig(folder + f'/Figure_Voltage_drops_Rpar={parasiticResistance[z]}_MW={memoryWindow[m]}.png')
            voltage_fig.show()


if Voltage_drops_error_plot:
    for m in range(memorySize):
        for z in range(np.size(parasiticResistance)):
            # Voltage drops comparison (Real vs Jeong vs DMR vs Gamma)
            # Vmax = np.max(relative_voltage_drops)
            # Create a figure with subplots side by side
            relative_voltage_fig = plt.figure(figsize=(15, 10))
            for index, model in enumerate(enabled_models):
                if model != reference_model:
                    relative_voltage_drops = np.abs((voltage_drops[:,:,z,m,reference_index] - voltage_drops[:,:,z,m,index]))/voltage_drops[:,:,z,m,reference_index]
                    # Plot heatmap for each model
                    plt.subplot(2, round((modelSize+1)/2), index+1)  # 1 row, 3 columns, 1st subplot
                    plt.imshow(relative_voltage_drops, cmap='hot', interpolation='nearest',vmax = 1 )
                    plt.colorbar(label='Voltage Drop error (V)')
                    plt.title('Voltage Drop error ('+ model +')')
                    plt.xlabel('Column Index (j)')
                    plt.ylabel('Row Index (m)')
            plt.tight_layout()
            plt.savefig(folder + f'/Figure_Relative_Voltage_drops_Rpar={parasiticResistance[z]}_MW={memoryWindow[m]}.png')
            relative_voltage_fig.show()


if Metric_plot:
    # 3D plots of the Metric
    X_plot, Y_plot = np.meshgrid(parasiticResistance, memoryWindow)
    x = X_plot.flatten()
    y = Y_plot.flatten()
    dx = dy = 0.5
    if np.size(parasiticResistance) > 1:
        dx = np.diff(parasiticResistance).min()
    if np.size(memoryWindow) > 1:
        dy = np.diff(memoryWindow).min()
    max_Metric = ((Metric.T).flatten()).max()
    num_models = len(enabled_models) - 1  # Exclude the reference model
    # Calculate optimal rows and columns for the subplots
    cols = int(np.ceil(np.sqrt(num_models)))
    rows = int(np.ceil(num_models / cols))
    # Create the subplots grid
    figure_Metric, axs = plt.subplots(rows, cols, subplot_kw={'projection': '3d'}, figsize=(15, 7))
    axs = np.array(axs).reshape(-1)  # Flatten to make indexing easier
    plot_index = 0
    for index, model in enumerate(enabled_models):
        if model != reference_model:
            dz = (Metric[:, :, index].T).flatten()
            axs[plot_index].bar3d(x, y, np.zeros_like(dz), dx, dy, dz, shade=True, color=colors[plot_index % len(colors)])
            axs[plot_index].set_title(model)
            axs[plot_index].set_xlabel('Parasitic Resistance (Ohm)')
            axs[plot_index].set_ylabel('Memory Window (Ohm)')
            axs[plot_index].set_zlabel(error_label)
            axs[plot_index].set_zlim(0, max_Metric)
            plot_index += 1
    # Hide any unused subplots
    for ax in axs[plot_index:]:
        ax.set_visible(False)
    plt.tight_layout()
    # Save and show the figure
    figure_Metric.savefig(folder + '/Figure_Metric_plot.png')
    plt.show()


if Mean_Metric_plot:
    print("mean Metric (each model):",np.mean(Metric, axis=0).mean(axis=0))
    # Total Metric histogram
    fig_Metric = plt.figure()
    # Prepare the data
    # Plot the histogram
    plt.bar(enabled_models[:-1], np.mean(Metric, axis=0).mean(axis=0), color=[colors[i % len(colors)] for i in range(len(enabled_models) - 1)])
    # Add labels and title
    plt.xlabel('Model')
    plt.ylabel(error_label)
    plt.title('Mean Score throught all simulations)\n' + array_size_string)
    plt.savefig(folder + '/Figure_Mean_Metric_plot.png')
    # Show the plot
    plt.show()


if Metric_vs_Rpar and memorySize==1:
    # Plotting
    plt.figure()
    for index, model in enumerate(enabled_models):
        if model != reference_model:
             plt.plot(parasiticResistance, Metric[:, 0, index], marker=markers[index % len(markers)], color=colors[index % len(colors)], label=model)
    # Log scale for y-axis as in the example image
    plt.xlabel("Line Resistance (Ω)")
    plt.ylabel(error_label)
    plt.title("Model Normalized Error vs Line Resistance\n"+array_size_string)
    plt.legend()
    plt.grid(True, which="both", linestyle='--', linewidth=0.5)
    plt.savefig(folder + '/Figure_error_vs_Rpar_plot.png')
    # Show the plot
    plt.show()

if Metric_vs_MW and parasiticSize==1:
    # Plotting
    plt.figure()
    for index, model in enumerate(enabled_models):
        if model != reference_model:
             plt.plot(memoryWindow, Metric[0, :, index], marker=markers[index % len(markers)], color=colors[index % len(colors)], label=model)
    # Log scale for y-axis as in the example image
    plt.xlabel("Memory Window (On/Off ratio)")
    plt.ylabel(error_label)
    plt.title("Model Normalized Error vs Memory Window\n"+array_size_string)
    plt.legend()
    plt.grid(True, which="both", linestyle='--', linewidth=0.5)
    plt.savefig(folder + '/Figure_error_vs_MW_plot.png')
    # Show the plot
    plt.show()

if print_table:
    model_labels = np.array(enabled_models)
    if Metric_type == 2:
        winning_indices = np.argmax(Metric, axis=-1)
    else:
        winning_indices = np.argmin(Metric, axis=-1)
    winning_models = model_labels[winning_indices]
    winning_df = pd.DataFrame(winning_models, index=parasiticResistance, columns=memoryWindow)
    # First save without formatting
    excel_path = folder_path + '/winning_models.xlsx'
    winning_df.to_excel(excel_path, index=True)
    # Define color mapping
    color_map = {
        'CrossSim': 'ADD8E6',     # light blue
        'Gamma': 'FFB6B6',        # light red
        'Gamma_acc': 'FFB6B6',    # light red
        'Gamma_acc_v2': 'FFB6B6',    # light red
        'Jeong': 'E0FFFF',        # light cyan
        'DMR': '90EE90'  ,        # light green
        'DMR_acc': '90EE90',       # light green
        'DMR_acc_v2': '90EE90'       # light green
    }
    # Open the saved file and add formatting
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    # Apply colors to cells (skip header row and index column)
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            model = cell.value
            if model in color_map:
                cell.fill = PatternFill(start_color=color_map[model],
                                      end_color=color_map[model],
                                      fill_type='solid')
    # Save the formatted file
    wb.save(excel_path)
    os.startfile(excel_path)

# Compute the winning model for each parasiticResistance vs. memoryWindow combination
if Winning_models_map:
    model_labels = np.array(enabled_models)
    # For error metrics (Metric_type=1 or 0), lower is better, so winning model = min metric
    # For Metric_type=2 (precision/time), higher is better, so winning model = max metric
    if Metric_type == 2:
        winning_indices = np.argmax(Metric, axis=-1)
    else:
        winning_indices = np.argmin(Metric, axis=-1)
    non_reference_models = enabled_models[:-1]
    if Metric_type == 2:
        winning_indices_nonref = np.argmax(Metric, axis=-1)
    else:
        winning_indices_nonref = np.argmin(Metric, axis=-1)
    
    # winning_indices_nonref now corresponds directly to indices in non_reference_models
    # Extract the metric values for these winning models
    # metric_winner: same shape as winning_indices_nonref, each entry is the winning model's metric value
    metric_winner = np.zeros_like(winning_indices_nonref, dtype=float)
    for z in range(parasiticSize):
        for m_ in range(memorySize):
            metric_winner[z, m_] = Metric[z, m_, winning_indices_nonref[z, m_]]
    # Identify the winning models' names
    winning_models_nonref = np.array(non_reference_models)[winning_indices_nonref]
    # Get min and max of the winning metric values for normalization
    metric_min = np.min(metric_winner)
    metric_max = np.max(metric_winner)
    if metric_min == metric_max:
        # Avoid division by zero if all metrics are equal
        metric_min = 0.0
        metric_max = 1.0
    # Prepare a 2D array of RGBA colors for the heatmap
    # We'll blend between white (low metric) and the model's base color (high metric)
    white = np.array([1.0, 1.0, 1.0])  # white color in RGB
    cell_colors = np.zeros((parasiticSize, memorySize, 3))
    # Create a dict mapping model to base RGB color
    # We already have 'colors' as a list of colors for each model in enabled_models
    # Convert colors to RGB if they are not already
    from matplotlib.colors import to_rgb
    model_color_map = {model: to_rgb(c) for model, c in zip(enabled_models, colors)}
    # Define a starting luminance that is not white but slightly darker gray
    light_gray = np.array([0.9, 0.9, 0.9]) 
    # Normalize and compute cell colors
    for z in range(parasiticSize):
        for m_ in range(memorySize):
            model = winning_models_nonref[z, m_]
            base_color = np.array(model_color_map[model])
            err_value = metric_winner[z, m_]
            # Normalize error to [0,1]
            scale = (err_value - metric_min) / (metric_max - metric_min)
            # Interpolate color: light_gray at low error, base_color at high error
            cell_color = light_gray * (1 - scale) + base_color * scale
            cell_colors[z, m_] = cell_color
    plt.figure(figsize=(6, 6))
    plt.imshow(cell_colors, origin="lower", aspect="auto",
            extent=[memoryWindow[0], memoryWindow[-1], parasiticResistance[0], parasiticResistance[-1]],
            interpolation="nearest")
    plt.xlabel("Memory Window")
    plt.ylabel("Parasitic Resistance")
    plt.title("Winning Models Map (Shaded by Error)")
    # Create legend patches with base colors (full saturation) of winning models
    winning_models_unique = [model for model in enabled_models if model in np.unique(winning_models_nonref)]
    legend_patches = [Patch(facecolor=model_color_map[model], label=model) for model in winning_models_unique]
    # Move the legend to the bottom center of the figure
    plt.legend(handles=legend_patches, loc="upper center", bbox_to_anchor=(0.5, -0.1), ncol=3)
    # Create a grayscale colormap from light_gray to black for the colorbar
    cmap_gray = LinearSegmentedColormap.from_list("gray_to_black", [light_gray, (0,0,0)], N=256)
    sm = ScalarMappable(cmap=cmap_gray, norm=plt.Normalize(vmin=metric_min, vmax=metric_max))
    sm.set_array([])
    # Get current figure and axes
    fig = plt.gcf()
    ax = plt.gca()
    # Add colorbar on the right side
    cbar = fig.colorbar(sm, ax=ax, fraction=0.046, pad=0.04)
    cbar.set_label("Relative error (%)")
    # Adjust layout to avoid overlap
    plt.tight_layout()
    plt.savefig(folder + '/Figure_Winning_Models_Map_Shaded.png')
    plt.show()



if scatter_plot:
    mean_Metric = np.mean(Metric, axis=0).mean(axis=0)
    variance_Metric = np.var(Metric, axis=0).mean(axis=0)  # Compute variance
    # Remove 'Ideal' if present, and handle arrays accordingly
    if "Ideal" in enabled_models:
        ideal_index = enabled_models.index("Ideal")
        # Remove 'Ideal' entry from models
        plot_models = [m for i, m in enumerate(enabled_models[:-1]) if i != ideal_index]
        # Remove the corresponding entries from arrays
        plot_times = np.delete(normalized_simulation_times[:-1], ideal_index)
        plot_Metric = np.delete(mean_Metric, ideal_index)
        plot_variance = np.delete(variance_Metric, ideal_index)
    else:
        # If Ideal not present, just use all models except the last one (assuming the last is a reference)
        plot_models = enabled_models[:-1]
        plot_times = simulation_times[:-1]
        plot_Metric = mean_Metric
        plot_variance = variance_Metric
    # Create a matching colors list for the plot_models
    # colors was initially defined to match enabled_models, so we rebuild a plot_colors array:
    model_to_color = {model: color for model, color in zip(enabled_models, colors)}
    plot_colors = [model_to_color[m] for m in plot_models]
    # Create a larger figure for better readability
    fig, ax = plt.subplots(figsize=(9, 6)) 
    # Identify CrossSim models among the plotted models
    crosssim_indices = [i for i, model in enumerate(plot_models) if model.startswith("CrossSim")]
    # If multiple CrossSim models are present, perform a regression on them
    if len(crosssim_indices) > 1:
        crosssim_times = plot_times[crosssim_indices].reshape(-1, 1)
        crosssim_values = plot_Metric[crosssim_indices]
        reg = LinearRegression().fit(np.log(crosssim_times), np.log(crosssim_values))  
        reg_line_x = np.linspace(crosssim_times.min(), crosssim_times.max(), 100).reshape(-1, 1)
        reg_line_y = np.exp(reg.predict(np.log(reg_line_x)))
        ax.plot(reg_line_x, reg_line_y, color='blue', linestyle='--', linewidth=2, label='CrossSim Regression')  # Increased linewidth
    # Plot scatter points with larger markers if needed
    scatter = ax.scatter(
        plot_times,
        plot_Metric,
        c=plot_colors,
        s=120,  # Marker size
        marker='o',
        edgecolor="black",
        linewidth=0.7
    )
    # Plot error bars
    for i, (x, y, var) in enumerate(zip(plot_times, plot_Metric, plot_variance)):
        ax.errorbar(
            x, y, yerr=np.sqrt(var), fmt='o', 
            color=plot_colors[i], capsize=5
        )
    # Add grid with adjusted properties if needed
    ax.grid(True, which="both", linestyle='--', linewidth=0.5, color="gray", alpha=0.7)
    # Create legend handles with larger font sizes
    legend_handles = []
    # Add handles for non-CrossSim models
    for i, model in enumerate(plot_models):
        if not model.startswith("CrossSim"):
            legend_handles.append(
                plt.Line2D([0], [0], marker='o', color='w', label=model,
                           markerfacecolor=plot_colors[i], markersize=10)
            )
    # Add a single handle for CrossSim if present
    if len(crosssim_indices) > 0:
        legend_handles.append(
            plt.Line2D([0], [0], color='blue', linestyle='--', label='CrossSim Regression')
        )
    # Add the legend with larger font size
    ax.legend(handles=legend_handles, title="Models",
              bbox_to_anchor=(1.05, 1), loc='upper left', borderaxespad=0.)
    # Set x-scale to log
    ax.set_xscale('log')
    # Set labels with larger font sizes
    ax.set_xlabel('Normalized Simulation Time (log scale)' if "Ideal" in enabled_models else 'Execution Time (seconds, log scale)', fontsize=14)
    ax.set_ylabel(error_label, fontsize=14)
    # Set title with larger font size
    ax.set_title('Scatter Plot' + ' (' + array_size_string +  ')\n')
    # Adjust tick parameters for larger font sizes
    ax.tick_params(axis='both', which='major')
    # Inset axis for low-error models
    inset_ax = inset_axes(ax, width="60%", height="60%", loc='upper right')
    threshold = 0.1  # Adjust as needed for "low-error" threshold
    low_error_indices = [i for i, metric in enumerate(plot_Metric) if metric < threshold]
    if len(low_error_indices) > 0:
        inset_times = plot_times[low_error_indices]
        inset_metrics = plot_Metric[low_error_indices]
        inset_variance = plot_variance[low_error_indices]
        inset_colors = [plot_colors[i] for i in low_error_indices]
        # Perform a linear regression on CrossSim models inside the inset if needed
        if len(crosssim_indices) > 1:
            inset_crosssim_times = plot_times[crosssim_indices].reshape(-1, 1)
            inset_crosssim_metrics = plot_Metric[crosssim_indices]
            inset_reg = LinearRegression().fit(np.log(inset_crosssim_times), np.log(inset_crosssim_metrics))
            inset_reg_line_x = np.linspace(inset_crosssim_times.min(), inset_crosssim_times.max(), 100).reshape(-1, 1)
            inset_reg_line_y = np.exp(inset_reg.predict(np.log(inset_reg_line_x)))
            inset_ax.plot(inset_reg_line_x, inset_reg_line_y, color='blue', linestyle='--', linewidth=2, label='Inset Regression')  # Increased linewidth
        # Plot scatter in inset
        inset_ax.scatter(
            inset_times,
            inset_metrics,
            c=inset_colors,
            s=60,  # Smaller marker size for inset
            marker='o',
            edgecolor="black",
            linewidth=0.7
        )
        # Plot error bars in inset
        for idx, (x, y, var) in enumerate(zip(inset_times, inset_metrics, inset_variance)):
            inset_ax.errorbar(
                x, y, yerr=np.sqrt(var), fmt='o', 
                color=inset_colors[idx],
                capsize=3
            )
        # Set scales to log
        inset_ax.set_xscale('log')
        inset_ax.set_yscale('log')
        # Add grid
        inset_ax.grid(True, which="both", linestyle='--', linewidth=0.5, alpha=0.7)
        # Set title with larger font size
        inset_ax.set_title("Logarithmic Zoom", fontsize=14)
        # Adjust tick parameters for larger font sizes
        inset_ax.tick_params(axis='both', which='major')
    # Adjust layout to accommodate legend
    plt.tight_layout()
    # Save the figure with higher DPI for better resolution
    plt.savefig(folder + '/Figure_Scatter_SimulationTimes_vs_error_with_inset.png', dpi=300)  # Increased DPI
    # Show the plot
    plt.show()


if Resistance_heatmap:
    # Heatmap resistance matrix
    fig, axes = plt.subplots(1, R.shape[2], figsize=(6 * R.shape[2], 4))
    axes = [axes] if R.shape[2] == 1 else axes
    for i, ax in enumerate(axes):
        # Plotting the heatmap using imshow
        cax = ax.imshow(R[0, 0, i], cmap='viridis', aspect='auto')
        ax.set_title(f"Resistance Array Variability Seed {i}")
        ax.set(xlabel='x position', ylabel='y position')
        # Adding the colorbar for each subplot
        fig.colorbar(cax, ax=ax, label='Resistance [R]', fraction=0.046, pad=0.04)
    plt.tight_layout()
    fig.savefig(folder + '/Heatmaps.png')
    plt.show()


if open_folder:
    os.startfile(folder_path)
